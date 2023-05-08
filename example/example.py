"""
Examples demonstrating the usage of `pyxltab`.
"""

import openpyxl
import pyxltab

WORKBOOK_FILENAME = r"example\\example.xlsx"


def get_cells_from_table():
    """
    Example demonstrating the `get_cells()` method at the table level.
    """

    openpyxl_book = openpyxl.load_workbook(WORKBOOK_FILENAME)
    book = pyxltab.attach(openpyxl_book)
    return book["Sheet1"]["Table1"].get_cells()


def get_cells_from_column():
    """
    Example demonstrating the usage of the `get_cells()` method at the column level.
    """
    openpyxl_book = openpyxl.load_workbook(WORKBOOK_FILENAME)
    book = pyxltab.attach(openpyxl_book)
    return book["Sheet1"]["Table1"]["Column1"].get_cells()


def get_cells_from_all_tables():
    """
    Example demonstrating the usage of `get_tables()`.
    """
    openpyxl_book = openpyxl.load_workbook(WORKBOOK_FILENAME)
    tables = pyxltab.get_tables(openpyxl_book)
    return [table.get_cells() for table in tables.values()]


if __name__ == "__main__":
    cells_in_table = get_cells_from_table()
    cells_in_column = get_cells_from_column()
    cells_in_all_tables = get_cells_from_all_tables()
